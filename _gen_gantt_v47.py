# -*- coding: utf-8 -*-
"""生成 LLM系統開發甘特圖_v4_7.xlsx
2026-04-27 更新：新增商機等級列、Phase 2 進度 40.5%、日期標題列
W1=4/01  W2=4/08  W3=4/15  W4=4/22（現在）  W5=4/29  W6=5/06  W7=5/13
W8=5/20  W9=5/27  W10=6/03  W11=6/10  W12=6/17  W13=6/24
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DST = r"d:\yujui\痛點需求地圖\prompt定版\LLM系統開發甘特圖_v4_7.xlsx"

# ── 顏色 ──────────────────────────────────────────────────────────
C_DONE    = "EF4444"   # 完成紅
C_RUN     = "F59E0B"   # 執行中黃/橙
C_PLAN    = "93C5FD"   # 規劃藍
C_NOW     = "FF185F"   # 當前週粉紅（W4）
C_BG_DARK = "0F172A"   # 深背景
C_HDR_BG  = "1E293B"   # 標題列背景
C_WHITE   = "FFFFFF"
C_GRAY    = "94A3B8"
C_LIGHT   = "F1F5F9"

# Phase 主色（用於 header 和 bar）
PHASE_COLORS = {
    "Phase 0":  "06B6D4",  # 青
    "Phase L":  "8B5CF6",  # 紫
    "Phase M":  "EC4899",  # 粉
    "Phase 1":  "F97316",  # 橘
    "商機等級": "EAB308",  # 黃
    "Phase 2":  "3B82F6",  # 藍
    "Phase 3":  "10B981",  # 綠
    "Phase 5":  "6B7280",  # 灰
}

def fill(hex6, fill_type="solid"):
    return PatternFill(fill_type=fill_type, fgColor="FF" + hex6)

def font(hex6=C_WHITE, bold=False, size=10, name="Microsoft JhengHei"):
    return Font(name=name, size=size, bold=bold, color="FF" + hex6)

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color="FF334155")
    return Border(left=s, right=s, top=s, bottom=s)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "甘特圖_v4.7"

# ── 欄寬 ──────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 11   # Phase
ws.column_dimensions["B"].width = 44   # 任務
ws.column_dimensions["C"].width = 8    # 類型
ws.column_dimensions["D"].width = 9    # 費用
for i in range(5, 18):                 # W1–W13
    ws.column_dimensions[get_column_letter(i)].width = 6

# ── 標題（R1）───────────────────────────────────────────────────
ws.row_dimensions[1].height = 28
c = ws.cell(1, 1, "業務日誌 LLM 系統開發甘特圖  v4.7  (2026-04-27)")
c.fill = fill(C_BG_DARK)
c.font = Font(name="Microsoft JhengHei", size=14, bold=True, color="FF" + "38BDF8")
ws.merge_cells("A1:Q1")
ws.cell(1, 1).alignment = align("left")

# ── 備註（R2）────────────────────────────────────────────────────
ws.row_dimensions[2].height = 16
note = ("v4.7 更新（2026-04-27）：新增「商機等級」三層 pipeline（關鍵詞→TF-IDF KNN→Batch LLM）；"
        "Phase 2 進度 40.5%（72,420行）；REPORT.html 互動式儀表板公開於 GitHub Pages")
c = ws.cell(2, 1, note)
c.fill = fill("1E3A52")
c.font = font("94A3B8", size=9)
ws.merge_cells("A2:Q2")
ws.cell(2, 1).alignment = align("left")

# ── 日期標題（R3）────────────────────────────────────────────────
ws.row_dimensions[3].height = 14
dates = ["", "", "", "", "4/01", "4/08", "4/15", "4/22▶", "4/29", "5/06",
         "5/13", "5/20", "5/27", "6/03", "6/10", "6/17", "6/24"]
for col, d in enumerate(dates, 1):
    c = ws.cell(3, col, d)
    c.fill = fill("0F172A")
    if "▶" in d:
        c.font = Font(name="Microsoft JhengHei", size=8, bold=True, color="FF" + C_NOW)
    else:
        c.font = font("64748B", size=8)
    c.alignment = align()

# ── 欄標題（R4）──────────────────────────────────────────────────
ws.row_dimensions[4].height = 20
headers = ["Phase", "任務 / 工作項目", "類型", "費用/成本",
           "W1", "W2", "W3", "W4", "W5", "W6", "W7", "W8", "W9", "W10", "W11", "W12", "W13"]
for col, h in enumerate(headers, 1):
    c = ws.cell(4, col, h)
    c.fill = fill(C_HDR_BG)
    c.font = font(C_LIGHT, bold=True, size=10)
    c.alignment = align()
    c.border = thin_border()

# ── 資料 ──────────────────────────────────────────────────────────
# bars: list of (col_offset_from_W1=0, length, status)
#   status: "done" | "run" | "plan"
# W1=col5, so col = 4 + w_num
#
# W1=1  W2=2  W3=3  W4=4  W5=5  W6=6  W7=7  W8=8  W9=9  W10=10  W11=11  W12=12  W13=13

ROWS = [
    # (phase, task, type, cost, bars_spec)
    # bars_spec: list of (w_start, w_end_inclusive, status)
    # status: "done" | "run" | "plan" | "now"（W4 = 當前週）
    ("Phase 0 ✓", "GCP 架構 + 問卷 Adapter + UEL schema",       "LLM",  "<$1",
     [(1,3,"done")]),
    (None, "  ✓ GCS bucket / Unified Event Log schema",           "GCP",  "$0",
     [(1,2,"done")]),
    (None, "  ✓ 問卷 Adapter → survey_signals.jsonl（82,105筆）", "LLM",  "$5–15",
     [(1,3,"done")]),
    (None, "  ✓ 三大信號萃取（Pain/Need/Insight）",              "LLM",  "$14–23",
     [(2,3,"done")]),

    ("Phase L ✓", "L1–L7 零人工分類 pipeline（1,802,590筆）",    "GCP",  "$30–46",
     [(1,4,"done")]),
    (None, "  ✓ Step 0：種子庫（2,375筆，KNN 73.67%）",          "LLM",  "$<1",
     [(1,3,"done")]),
    (None, "  ✓ Step 1：規則快篩（867,023筆 HIGH，20.9%）",       "GCP",  "$0",
     [(2,3,"done")]),
    (None, "  ✓ Step 2：Embedding+KNN（935,570筆 HIGH 100%）",   "GCP",  "$15–20",
     [(3,4,"done")]),
    (None, "  ✓ Step 3：SC LLM（14筆）→ phase_l_final 180萬",   "LLM",  "$<1",
     [(4,4,"done")]),

    ("Phase M ✓", "分層聚類 + Map 歸納 + 熱路徑",                "GCP",  "$13–23",
     [(4,5,"done")]),
    (None, "  ✓ BigQuery ML KMeans(k=7) + PCA",                  "GCP",  "$3–8",
     [(4,5,"done")]),
    (None, "  ✓ 7 種法人類型命名 + 熱路徑 Top-30",               "LLM",  "$10–15",
     [(5,5,"done")]),
    (None, "  ✓ REPORT.html 互動式儀表板（GitHub Pages）",        "—",    "$0",
     [(5,5,"done")]),

    ("Phase 1 ✓", "法人歷程標籤（206,817家，8大類）",             "LLM",  "$170–370",
     [(3,5,"done")]),
    (None, "  ✓ Stage 1/2/3 並行 LLM（RPM=30，6 Workers）",      "LLM",  "$—",
     [(3,5,"done")]),
    (None, "  ✓ 8維屬性：外商/家族/規模/貿易國/品牌…",           "LLM",  "$—",
     [(4,5,"done")]),

    ("商機等級", "LeadInfo 銷售漏斗（756,989筆 E/D/C2/C1/B/A）", "LLM",  "低",
     [(4,4,"now"), (5,7,"run")]),
    (None, "  Layer 1：關鍵詞快篩（~70%，零API）",               "規則", "$0",
     [(4,4,"now"), (5,6,"run")]),
    (None, "  Layer 2：TF-IDF KNN k=5（~25%，零API）",           "ML",   "$0",
     [(4,4,"now"), (5,6,"run")]),
    (None, "  Layer 3：Batch LLM 10筆/次（~5%）",                "LLM",  "極低",
     [(5,7,"run")]),

    ("Phase 2",  "日報深度標籤（L1–L7 × 3欄，~178,790行）",      "LLM",  "$—",
     [(4,4,"now"), (5,7,"run")]),
    (None, "  🔄 40.5% 完成（72,420行），RESUME=True，6 Workers","LLM",  "$—",
     [(4,4,"now"), (5,6,"run")]),
    (None, "  L1–L4 結構化（pain/role/goal/issue）",              "LLM",  "$—",
     [(5,6,"run")]),
    (None, "  L5–L7 結構化（evaluate/direction/outcome）",        "LLM",  "$—",
     [(6,7,"run")]),

    ("Phase 3",  "痛需熱圖 + 三輸出邏輯",                        "—",    "$—",
     [(8,10,"plan")]),
    (None, "  JOIN：Phase2 × 商機等級 × PhaseM by company_id",   "—",    "$—",
     [(8,8,"plan")]),
    (None, "  痛需熱圖：法人類型 × L1痛點，色=B+A比例",          "—",    "$—",
     [(8,9,"plan")]),
    (None, "  Output 1：推薦業務問項（空白議題補問）",            "—",    "$—",
     [(9,9,"plan")]),
    (None, "  Output 2：開發方案卡（heat_score≥0.5）",            "—",    "$—",
     [(9,10,"plan")]),
    (None, "  Output 3：機會卡（B+A 高密度法人×痛點）",          "—",    "$—",
     [(10,10,"plan")]),

    ("Phase 5",  "擴展新資料源（客服/會議/外部）",                "—",    "$—",
     [(11,13,"plan")]),
    (None, "  客服記錄 Adapter",                                  "—",    "$—",
     [(11,11,"plan")]),
    (None, "  會議記錄 / 外部資料 Adapter",                       "—",    "$—",
     [(12,12,"plan")]),
]

# ── 寫入資料列 ────────────────────────────────────────────────────
def bar_fill(phase_key, status):
    if status == "done":
        return fill(C_DONE)
    elif status == "now":
        return fill(C_NOW)
    elif status == "run":
        return fill(C_RUN)
    elif status == "plan":
        pc = PHASE_COLORS.get(phase_key, "93C5FD")
        return fill(pc + "66"[:2] if len(pc) == 6 else "93C5FD")  # lighter
    return None

EMPTY_FILL  = fill("1E293B")   # W 欄空格
PHASE_ROW_ROW_H = 22
TASK_ROW_H      = 16

current_phase = None
row = 5

for (phase, task, typ, cost, bars) in ROWS:
    is_phase_hdr = phase is not None
    if is_phase_hdr:
        current_phase = phase.replace(" ✓", "").strip()
        ws.row_dimensions[row].height = PHASE_ROW_ROW_H
    else:
        ws.row_dimensions[row].height = TASK_ROW_H

    phase_key = current_phase

    # A: Phase 名稱
    pc = PHASE_COLORS.get(phase_key, "64748B")
    ca = ws.cell(row, 1, phase if is_phase_hdr else "")
    if is_phase_hdr:
        ca.fill = fill(pc)
        ca.font = Font(name="Microsoft JhengHei", size=10, bold=True, color="FF" + C_WHITE)
    else:
        ca.fill = fill("0F172A")
        ca.font = font("334155", size=9)
    ca.alignment = align("center")

    # B: 任務
    cb = ws.cell(row, 2, task)
    if is_phase_hdr:
        cb.fill = fill(pc)
        cb.font = Font(name="Microsoft JhengHei", size=10, bold=True, color="FF" + C_WHITE)
    else:
        cb.fill = fill("0F172A")
        cb.font = font("CBD5E1", size=9)
    cb.alignment = align("left", wrap=True)

    # C: 類型
    cc = ws.cell(row, 3, typ)
    cc.fill = fill("0F172A" if not is_phase_hdr else pc)
    cc.font = font(C_WHITE if is_phase_hdr else "64748B", size=9)
    cc.alignment = align()

    # D: 費用
    cd = ws.cell(row, 4, cost)
    cd.fill = fill("0F172A" if not is_phase_hdr else pc)
    cd.font = font(C_WHITE if is_phase_hdr else "64748B", size=9)
    cd.alignment = align()

    # W1–W13 bars
    bar_map = {}  # w_num → status
    for (ws_, we, st) in bars:
        for wn in range(ws_, we + 1):
            bar_map[wn] = st

    for wn in range(1, 14):
        col = 4 + wn
        cell = ws.cell(row, col, "")
        if wn in bar_map:
            st = bar_map[wn]
            if st == "done":
                cell.fill = fill(C_DONE)
            elif st == "now":
                cell.fill = fill(C_NOW)
            elif st == "run":
                cell.fill = fill(C_RUN)
            elif st == "plan":
                pcolor = PHASE_COLORS.get(phase_key, "93C5FD")
                cell.fill = fill(pcolor)
                cell.font = font(pcolor, size=8)
        else:
            cell.fill = EMPTY_FILL
        cell.alignment = align()
        cell.border = thin_border()

    row += 1

# ── 圖例（最後兩行）──────────────────────────────────────────────
ws.row_dimensions[row].height = 14
legend_items = [
    (C_DONE, "已完成（紅）"),
    (C_NOW,  "當前週（W4，4/22）"),
    (C_RUN,  "執行中"),
    "93C5FD:規劃中",
]
ws.cell(row, 1, "圖例").fill = fill(C_HDR_BG)
ws.cell(row, 1).font = font(C_GRAY, bold=True, size=9)
col_l = 2
for item in legend_items:
    if isinstance(item, tuple):
        clr, label = item
        ws.cell(row, col_l, "").fill = fill(clr)
        ws.cell(row, col_l + 1, label).fill = fill(C_HDR_BG)
        ws.cell(row, col_l + 1).font = font(C_LIGHT, size=9)
    else:
        clr, label = item.split(":")
        ws.cell(row, col_l, "").fill = fill(clr)
        ws.cell(row, col_l + 1, label).fill = fill(C_HDR_BG)
        ws.cell(row, col_l + 1).font = font(C_LIGHT, size=9)
    col_l += 2

# ── 凍結窗格 ─────────────────────────────────────────────────────
ws.freeze_panes = "E5"

# ── 儲存 ──────────────────────────────────────────────────────────
wb.save(DST)
print(f"OK: {DST}")
print(f"共 {row - 5} 筆資料列")
